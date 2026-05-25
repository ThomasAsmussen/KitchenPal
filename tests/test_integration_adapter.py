import shutil
import tempfile
from datetime import date

import openpyxl

from kitchenpal import constants
from kitchenpal.sheets_service import SheetsService


class OpenpyxlWritableAdapter:
    def __init__(self, workbook_path, sheet_name):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path)
        self.ws = self.wb[sheet_name]

    def _a1_to_rc(self, a1):
        import re

        m = re.match(r"([A-Z]+)(\d+)", a1)
        col_letters = m.group(1)
        row = int(m.group(2))
        col = 0
        for ch in col_letters:
            col = col * 26 + (ord(ch) - 64)
        return row, col

    def batch_get(self, ranges):
        out = []
        for r in ranges:
            start, end = r.split(":")
            r1, c1 = self._a1_to_rc(start)
            r2, c2 = self._a1_to_rc(end)
            vals = []
            for row in range(r1, r2 + 1):
                rowvals = []
                for col in range(c1, c2 + 1):
                    rowvals.append(self.ws.cell(row=row, column=col).value)
                vals.append(rowvals)
            out.append(vals)
        return out

    def cell(self, row, col):
        from types import SimpleNamespace

        return SimpleNamespace(value=self.ws.cell(row=row, column=col).value)

    def update_cell(self, row, col, value):
        self.ws.cell(row=row, column=col).value = value

    def update_acell(self, a1_ref, value):
        r, c = self._a1_to_rc(a1_ref)
        self.update_cell(r, c, value)

    def batch_update(self, updates):
        for upd in updates:
            rng = upd["range"]
            values = upd.get("values", [])
            start, end = rng.split(":") if ":" in rng else (rng, rng)
            r1, c1 = self._a1_to_rc(start)
            r2, c2 = self._a1_to_rc(end)
            # write values row by row
            for i, rowvals in enumerate(values):
                for j, val in enumerate(rowvals):
                    self.ws.cell(row=r1 + i, column=c1 + j).value = val

    def add_rows(self, n):
        self.ws.insert_rows(self.ws.max_row + 1, n)

    @property
    def row_count(self):
        return self.ws.max_row

    def get_all_values(self):
        rows = []
        for r in range(1, self.ws.max_row + 1):
            rowvals = [self.ws.cell(row=r, column=c).value for c in range(1, self.ws.max_column + 1)]
            rows.append(rowvals)
        return rows

    def clear(self):
        for r in range(1, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):
                self.ws.cell(row=r, column=c).value = None

    def save(self, path=None):
        path = path or self.workbook_path
        self.wb.save(path)


class FakeSpreadsheetAdapter:
    def __init__(self, adapter):
        self._adapter = adapter

    def worksheet(self, name):
        return self._adapter


def _copy_test_sheet():
    src = "test_sheet.xlsx"
    tmp = tempfile.NamedTemporaryFile(prefix="kitchenpal_test_", suffix=".xlsx", delete=False)
    tmp.close()
    shutil.copy(src, tmp.name)
    return tmp.name


def test_integration_add_purchase_and_drinks():
    tmp = _copy_test_sheet()
    adapter = OpenpyxlWritableAdapter(tmp, "Maj 2026")
    fake_spreadsheet = FakeSpreadsheetAdapter(adapter)

    # build minimal SheetsService and inject fake spreadsheet
    service = SheetsService.__new__(SheetsService)
    service._spreadsheet = fake_spreadsheet

    # Add a purchase - should write AC..AE and AG on first empty lookup row
    service.add_purchase("Maj 2026", 352, date(2026, 5, 24), "Banankage", 42.0)

    # verify AC..AE and AG on the sheet by reading the adapter across a larger range
    rng = adapter.batch_get([f"{constants.PURCHASE_INSERT_START_COLUMN}2:{constants.PURCHASE_AMOUNT_COLUMN}300"])[0]
    found = False
    for i, row in enumerate(rng, start=2):
        # match room 352 and item substring 'Banankage'
        try:
            room_cell = row[0]
            item_cell = row[2]
        except Exception:
            continue
        if room_cell is None:
            continue
        if str(room_cell).startswith("352") and item_cell and "Banankage" in str(item_cell):
            found = True
            break
    assert found, "Purchase not written to expected range"

    # Add drinks for room 346 - should increment value in PERSONAL_ACCOUNT_BEER_COLUMN
    # read current value
    # find account row via helper
    account_row = service._find_account_row_in_kovs(adapter, "346")
    assert account_row is not None
    before_beer = adapter.cell(account_row, constants.PERSONAL_ACCOUNT_BEER_COLUMN).value or 0
    before_wine = adapter.cell(account_row, constants.PERSONAL_ACCOUNT_WINE_COLUMN).value or 0

    service.add_drinks("Maj 2026", 346, 2, 1)

    after_beer = adapter.cell(account_row, constants.PERSONAL_ACCOUNT_BEER_COLUMN).value
    after_wine = adapter.cell(account_row, constants.PERSONAL_ACCOUNT_WINE_COLUMN).value

    assert int(after_beer) == int(before_beer) + 2
    assert int(after_wine) == int(before_wine) + 1

    # cleanup
    adapter.save(tmp)


def test_integration_add_three_wines_increases_ak_row_three_by_three():
    tmp = _copy_test_sheet()
    adapter = OpenpyxlWritableAdapter(tmp, "Maj 2026")
    fake_spreadsheet = FakeSpreadsheetAdapter(adapter)

    service = SheetsService.__new__(SheetsService)
    service._spreadsheet = fake_spreadsheet

    account_row = service._find_account_row_in_kovs(adapter, "346")
    assert account_row == 3

    # Seed AK3 with a known baseline so we can verify the exact delta.
    baseline = 4
    adapter.update_cell(3, constants.PERSONAL_ACCOUNT_WINE_COLUMN, baseline)

    service.add_drinks("Maj 2026", 346, 0, 3)

    after_wine = adapter.cell(3, constants.PERSONAL_ACCOUNT_WINE_COLUMN).value
    assert int(after_wine) == baseline + 3

    adapter.save(tmp)
